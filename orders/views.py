from rest_framework import generics, permissions
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
import io
import os

from .models import Order
from .serializers import OrderSerializer


@api_view(['GET'])
@permission_classes([permissions.AllowAny])
def order_detail_public(request, pk):
    order = get_object_or_404(Order, pk=pk)
    serializer = OrderSerializer(order)
    return Response(serializer.data)


class OrderListCreateView(generics.ListCreateAPIView):
    serializer_class = OrderSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return Order.objects.filter(user=self.request.user).prefetch_related('items')

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


class OrderRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = OrderSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return Order.objects.filter(user=self.request.user).prefetch_related('items')


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def export_excel(request, pk):
    order = get_object_or_404(Order, pk=pk, user=request.user)
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, Alignment, PatternFill, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    import urllib.request

    wb = Workbook()
    ws = wb.active
    ws.title = "Зayavka"

    # Column widths
    col_widths = [5, 22, 18, 16, 38, 16, 8, 16, 12, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Row heights
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20

    header_font = Font(name='Times New Roman', bold=True, size=11)
    normal_font = Font(name='Times New Roman', size=10)
    red_font = Font(name='Times New Roman', size=10, color='FF0000', bold=True)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)

    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header rows
    ws.merge_cells('A1:C1')
    ws['A1'] = f"Заказчик:   {order.client_name}"
    ws['A1'].font = header_font

    ws.merge_cells('D1:J1')
    ws['D1'] = f"Tel: {order.client_phone}"
    ws['D1'].font = header_font

    ws.merge_cells('A2:J2')
    ws['A2'] = f"Дата заявки:  {order.created_at.strftime('%d, %m %Y')}"
    ws['A2'].font = header_font

    # Table header row
    headers = ['№', 'Наименование', 'Вид', 'Габариты', 'Краткая характеристика',
               'Стоимость', 'Кол. во', 'Сумма', 'НДС 12%', 'сумма с учётом НДС 12%']

    green_fill = PatternFill(fill_type='solid', fgColor='C6EFCE')
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = Font(name='Times New Roman', bold=True, size=10)
        cell.alignment = center
        cell.border = border
        cell.fill = green_fill
    ws.row_dimensions[3].height = 30

    # Data rows
    for row_num, item in enumerate(order.items.all(), 1):
        row = 3 + row_num
        ws.row_dimensions[row].height = 70

        data = [
            row_num,
            item.name,
            '',  # image placeholder
            item.dimensions,
            item.description,
            int(item.price),
            item.quantity,
            int(item.total_price),
            int(item.vat_amount),
            int(item.grand_total),
        ]

        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = border
            cell.alignment = center
            if col_idx == 2:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            if col_idx == 5:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            if col_idx in (4,):
                cell.font = red_font
            else:
                cell.font = normal_font

        # Try to embed the product image
        if item.image_url:
            try:
                from openpyxl.drawing.image import Image as XLImage
                img_url = item.image_url
                if img_url.startswith('/'):
                    img_url = request.build_absolute_uri(img_url)
                req = urllib.request.Request(img_url, headers={'User-Agent': 'Mozilla/5.0'})
                img_data = urllib.request.urlopen(req, timeout=5).read()
                img_stream = io.BytesIO(img_data)
                xl_img = XLImage(img_stream)
                xl_img.width = 110
                xl_img.height = 80
                cell_addr = f"C{row}"
                ws.add_image(xl_img, cell_addr)
            except Exception:
                pass

    # Totals rows
    last_data_row = 3 + order.items.count()
    total_row = last_data_row + 1
    vat_row = last_data_row + 2
    grand_row = last_data_row + 3

    ws.row_dimensions[total_row].height = 20
    ws.row_dimensions[vat_row].height = 20
    ws.row_dimensions[grand_row].height = 20

    for r, label, val in [
        (total_row, 'Итого:', int(order.total_sum)),
        (vat_row, 'НДС 12%:', int(order.vat_sum)),
        (grand_row, 'Итого с НДС:', int(order.grand_total)),
    ]:
        ws.merge_cells(f'A{r}:G{r}')
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = Font(name='Times New Roman', bold=True, size=11)
        lc.alignment = Alignment(horizontal='right', vertical='center')
        lc.border = border

        vc = ws.cell(row=r, column=8, value=val)
        vc.font = Font(name='Times New Roman', bold=True, size=11)
        vc.alignment = center
        vc.border = border

        for c in range(9, 11):
            ws.cell(row=r, column=c).border = border

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"zayavka_{order.pk}_{order.created_at.strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def export_pdf(request, pk):
    order = get_object_or_404(Order, pk=pk, user=request.user)
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import urllib.request

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=10 * mm,
        leftMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )

    styles = getSampleStyleSheet()
    bold_style = ParagraphStyle('Bold', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=11)
    normal_style = ParagraphStyle('Norm', parent=styles['Normal'], fontName='Helvetica', fontSize=9)
    center_style = ParagraphStyle('Center', parent=styles['Normal'], fontName='Helvetica', fontSize=9, alignment=1)
    red_style = ParagraphStyle('Red', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=9,
                               textColor=colors.red, alignment=1)
    wrap_style = ParagraphStyle('Wrap', parent=styles['Normal'], fontName='Helvetica', fontSize=8,
                                leading=10, alignment=1)

    story = []

    story.append(Paragraph(f"<b>Заказчик: {order.client_name}</b>   Tel: {order.client_phone}", bold_style))
    story.append(Spacer(1, 4 * mm))
    story.append(Paragraph(f"<b>Дата заявки: {order.created_at.strftime('%d.%m.%Y')}</b>", bold_style))
    story.append(Spacer(1, 6 * mm))

    header_bg = colors.HexColor('#C6EFCE')

    col_headers = [
        Paragraph('<b>№</b>', center_style),
        Paragraph('<b>Наименование</b>', center_style),
        Paragraph('<b>Вид</b>', center_style),
        Paragraph('<b>Габариты</b>', center_style),
        Paragraph('<b>Краткая характеристика</b>', center_style),
        Paragraph('<b>Стоимость</b>', center_style),
        Paragraph('<b>Кол.во</b>', center_style),
        Paragraph('<b>Сумма</b>', center_style),
        Paragraph('<b>НДС 12%</b>', center_style),
        Paragraph('<b>Сумма с НДС</b>', center_style),
    ]

    table_data = [col_headers]

    for idx, item in enumerate(order.items.all(), 1):
        img_cell = ''
        if item.image_url:
            try:
                img_url = item.image_url
                if img_url.startswith('/'):
                    img_url = request.build_absolute_uri(img_url)
                req = urllib.request.Request(img_url, headers={'User-Agent': 'Mozilla/5.0'})
                img_bytes = urllib.request.urlopen(req, timeout=5).read()
                img_stream = io.BytesIO(img_bytes)
                rl_img = RLImage(img_stream, width=28 * mm, height=22 * mm)
                img_cell = rl_img
            except Exception:
                img_cell = ''

        row = [
            Paragraph(str(idx), center_style),
            Paragraph(item.name, wrap_style),
            img_cell,
            Paragraph(f'<font color="red"><b>{item.dimensions}</b></font>', center_style),
            Paragraph(item.description, wrap_style),
            Paragraph(f"{int(item.price):,}".replace(',', ' '), center_style),
            Paragraph(str(item.quantity), center_style),
            Paragraph(f"{int(item.total_price):,}".replace(',', ' '), center_style),
            Paragraph(f"{int(item.vat_amount):,}".replace(',', ' '), center_style),
            Paragraph(f"{int(item.grand_total):,}".replace(',', ' '), center_style),
        ]
        table_data.append(row)

    # Totals
    empty7 = ['', '', '', '', '', '', '']
    table_data.append(
        [Paragraph('<b>Итого:</b>', bold_style), '', '', '', '', '', '',
         Paragraph(f"<b>{int(order.total_sum):,}</b>".replace(',', ' '), center_style), '', '']
    )
    table_data.append(
        [Paragraph('<b>НДС 12%:</b>', bold_style), '', '', '', '', '', '',
         Paragraph(f"<b>{int(order.vat_sum):,}</b>".replace(',', ' '), center_style), '', '']
    )
    table_data.append(
        [Paragraph('<b>Итого с НДС:</b>', bold_style), '', '', '', '', '', '',
         Paragraph(f"<b>{int(order.grand_total):,}</b>".replace(',', ' '), center_style), '', '']
    )

    col_widths_pdf = [
        10 * mm, 38 * mm, 32 * mm, 22 * mm, 60 * mm,
        25 * mm, 14 * mm, 25 * mm, 20 * mm, 30 * mm
    ]

    t = Table(table_data, colWidths=col_widths_pdf, repeatRows=1)

    num_data_rows = order.items.count()
    row_heights = [12 * mm] + [28 * mm] * num_data_rows + [8 * mm, 8 * mm, 8 * mm]

    t._argH = row_heights

    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), header_bg),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, num_data_rows), 0.5, colors.black),
        ('BOX', (0, 0), (-1, num_data_rows), 1, colors.black),
        # Totals rows
        ('SPAN', (0, num_data_rows + 1), (6, num_data_rows + 1)),
        ('SPAN', (0, num_data_rows + 2), (6, num_data_rows + 2)),
        ('SPAN', (0, num_data_rows + 3), (6, num_data_rows + 3)),
        ('ALIGN', (0, num_data_rows + 1), (6, num_data_rows + 3), 'RIGHT'),
        ('BOX', (0, num_data_rows + 1), (-1, num_data_rows + 3), 1, colors.black),
        ('GRID', (0, num_data_rows + 1), (-1, num_data_rows + 3), 0.5, colors.black),
    ])
    t.setStyle(style)

    story.append(t)
    doc.build(story)

    buffer.seek(0)
    filename = f"zayavka_{order.pk}_{order.created_at.strftime('%Y%m%d')}.pdf"
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response



